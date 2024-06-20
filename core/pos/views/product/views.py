import json
from datetime import datetime
from io import BytesIO


import xlsxwriter
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse_lazy
from django.views.generic import CreateView, UpdateView, DeleteView, TemplateView
from django.views.generic.base import View
from openpyxl import load_workbook

from core.pos.forms import ProductForm, Product, Category
from core.pos.models import Inventory
from core.security.mixins import GroupPermissionMixin


class ProductListView(GroupPermissionMixin, TemplateView):
    template_name = 'product/list.html'
    permission_required = 'view_product'

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'search':
                data = []
                for i in Product.objects.filter():
                    data.append(i.toJSON())
            elif action == 'search_inventory':
                data = []
                for i in Inventory.objects.filter(product_id=request.POST['id'], active=True).order_by('expiration_date', 'date_joined'):
                    data.append(i.toJSON())
            elif action == 'upload_excel':
                with transaction.atomic():
                    archive = request.FILES['archive']
                    workbook = load_workbook(filename=archive, data_only=True)
                    wb = workbook[workbook.sheetnames[0]]
                    for cell in wb.iter_rows(min_row=2, max_row=wb.max_row):
                        row = cell[0].row
                        id = int(wb.cell(row=row, column=1).value)
                        product = Product.objects.filter(id=id).first()
                        if product is None:
                            product = Product()
                        product.name = wb.cell(row=row, column=2).value
                        product.code = str(wb.cell(row=row, column=3).value)
                        product.category = Category.objects.get_or_create(name=wb.cell(row=row, column=4).value)[0]
                        product.price = float(wb.cell(row=row, column=5).value)
                        product.pvp = float(wb.cell(row=row, column=6).value)
                        product.stock = int(wb.cell(row=row, column=7).value)
                        product.inventoried = wb.cell(row=row, column=8).value.lower() == 'si'
                        product.with_tax = wb.cell(row=row, column=9).value.lower() == 'si'
                        product.save()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Productos'
        context['create_url'] = reverse_lazy('product_create')
        return context


class ProductCreateView(GroupPermissionMixin, CreateView):
    model = Product
    template_name = 'product/create.html'
    form_class = ProductForm
    success_url = reverse_lazy('product_list')
    permission_required = 'add_product'

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'add':
                form = self.get_form()
                if form.is_valid():
                    product = form.save()
                    price_list = [{**item, 'gross_price': product.calculate_gross_price(float(item['net_price'])), 'net_price': float(item['net_price'])} for item in json.loads(request.POST['price_list'])]
                    product.price_list = price_list
                    product.edit()
                else:
                    data['error'] = form.errors
            elif action == 'validate_data':
                data = {'valid': True}
                queryset = Product.objects.all()
                pattern = request.POST['pattern']
                if pattern == 'name':
                    name = request.POST['name'].strip()
                    category = request.POST['category']
                    if len(category):
                        data['valid'] = not queryset.filter(name__iexact=name, category_id=category).exists()
                elif pattern == 'code':
                    data['valid'] = not queryset.filter(code__iexact=request.POST['code']).exists()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['title'] = 'Nuevo registro de un Producto'
        context['list_url'] = self.success_url
        context['action'] = 'add'
        context['prices'] = []
        return context


class ProductUpdateView(GroupPermissionMixin, UpdateView):
    model = Product
    template_name = 'product/create.html'
    form_class = ProductForm
    success_url = reverse_lazy('product_list')
    permission_required = 'change_product'

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'edit':
                form = self.get_form()
                if form.is_valid():
                    product = form.save()
                    price_list = [{**item, 'gross_price': product.calculate_gross_price(float(item['net_price'])), 'net_price': float(item['net_price'])} for item in json.loads(request.POST['price_list'])]
                    print(price_list)
                    product.price_list = price_list
                    product.edit()
                else:
                    data['error'] = form.errors
            elif action == 'validate_data':
                data = {'valid': True}
                id = self.get_object().id
                queryset = Product.objects.all().exclude(id=id)
                pattern = request.POST['pattern']
                if pattern == 'name':
                    name = request.POST['name'].strip()
                    category = request.POST['category']
                    if len(category):
                        data['valid'] = not queryset.filter(name__iexact=name, category_id=category).exists()
                elif pattern == 'code':
                    data['valid'] = not queryset.filter(code__iexact=request.POST['code']).exists()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['title'] = 'Edición de un Producto'
        context['list_url'] = self.success_url
        context['action'] = 'edit'
        context['prices'] = json.dumps(self.object.price_list) if self.object.price_list else []
        return context


class ProductDeleteView(GroupPermissionMixin, DeleteView):
    model = Product
    template_name = 'delete.html'
    success_url = reverse_lazy('product_list')
    permission_required = 'delete_product'

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            self.get_object().delete()
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Notificación de eliminación'
        context['list_url'] = self.success_url
        return context


class ProductStockAdjustmentView(GroupPermissionMixin, TemplateView):
    template_name = 'product/stock_adjustment.html'
    permission_required = 'adjust_product_stock'

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'search_product':
                data = []
                ids = json.loads(request.POST['ids'])
                term = request.POST['term']
                queryset = Inventory.objects.filter(active=True).exclude(id__in=ids).order_by('product__name')
                if len(term):
                    queryset = queryset.filter(Q(product__name__icontains=term) | Q(product__code__icontains=term))
                    queryset = queryset[0:10]
                for i in queryset:
                    item = i.toJSON()
                    item['product'] = i.product.toJSON()
                    item['value'] = i.get_full_name()
                    data.append(item)
            elif action == 'create':
                with transaction.atomic():
                    for i in json.loads(request.POST['products']):
                        inventory = Inventory.objects.get(pk=i['id'])
                        inventory.quantity = int(i['newstock'])
                        inventory.saldo = int(i['newstock'])
                        inventory.save()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Ajuste de Stock de Productos'
        return context


class ProductExportExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        try:
            headers = {'Id': 15, 'Nombre': 75, 'Código': 20, 'Categoría': 20, 'Precio de Compra': 20, 'Precio de Venta': 20, 'Stock': 10, '¿Es inventariado?': 15, '¿Se cobra impuesto?': 15}
            output = BytesIO()
            workbook = xlsxwriter.Workbook(output)
            worksheet = workbook.add_worksheet('productos')
            cell_format = workbook.add_format({'bold': True, 'align': 'center', 'border': 1})
            row_format = workbook.add_format({'align': 'center', 'border': 1})
            index = 0
            for name, width in headers.items():
                worksheet.set_column(first_col=index, last_col=index, width=width)
                worksheet.write(0, index, name, cell_format)
                index += 1
            row = 1
            for product in Product.objects.filter().order_by('id'):
                worksheet.write(row, 0, product.id, row_format)
                worksheet.write(row, 1, product.name, row_format)
                worksheet.write(row, 2, product.code, row_format)
                worksheet.write(row, 3, product.category.name, row_format)
                worksheet.write(row, 4, f'{product.price:.2f}', row_format)
                worksheet.write(row, 5, f'{product.pvp:.2f}', row_format)
                worksheet.write(row, 6, product.stock, row_format)
                worksheet.write(row, 7, 'Si' if product.inventoried else 'No', row_format)
                worksheet.write(row, 8, 'Si' if product.with_tax else 'No', row_format)
                row += 1
            workbook.close()
            output.seek(0)
            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f"attachment; filename=PRODUCTOS_{datetime.now().date().strftime('%d_%m_%Y')}.xlsx"
            return response
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('product_list'))
